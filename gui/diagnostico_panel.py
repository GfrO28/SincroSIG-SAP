# gui/diagnostico_panel.py
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import date

import ttkbootstrap as tb
import pandas as pd
import pyperclip

from config.utils import set_window_icon
from config import uom_dict as _uom
from gui.widgets.calendar_widget import CampoFecha
from modules.diagnostico_queries import (
    construir_query_config,
    construir_query_negativos,
    construir_query_saldos_sap,
    construir_query_uom_sap,
)

# ── Normalización de columnas SAP (OITM) ────────────────────────────────────
_RENAME_OITM = {
    # ItemCode
    'ItemCode': 'ItemCode', 'Código de artículo': 'ItemCode',
    'Número de artículo': 'ItemCode', 'Artículo': 'ItemCode',
    'Código': 'ItemCode', 'Cod. Artículo': 'ItemCode',
    # ItemName
    'ItemName': 'ItemName', 'Nombre del artículo': 'ItemName',
    'Nombre de artículo': 'ItemName', 'Nombre artículo': 'ItemName',
    'Nombre': 'ItemName', 'Descripción': 'ItemName',
    'Descripción del artículo': 'ItemName',
    # UOM inventario
    'InvntryUom': 'UOM_Inv', 'UOM_Inventario': 'UOM_Inv',
    'UOM de inventario': 'UOM_Inv', 'UM inventario': 'UOM_Inv',
    'Ud. de medida inventario': 'UOM_Inv',
    # UOM compras
    'BuyUnitMsr': 'UOM_Com', 'UOM_Compras': 'UOM_Com',
    'UOM de compras': 'UOM_Com', 'UM compras': 'UOM_Com',
    'Ud. de medida compra': 'UOM_Com',
    # Factor compras
    'NumInBuy': 'Factor_Com', 'Factor_Compras': 'Factor_Com',
    'Artículos por unidad de compra': 'Factor_Com',
    # UOM ventas
    'SalUnitMsr': 'UOM_Ven', 'UOM_Ventas': 'UOM_Ven',
    'UOM de ventas': 'UOM_Ven', 'UM ventas': 'UOM_Ven',
    'Ud. de medida venta': 'UOM_Ven',
    # Factor ventas
    'NumInSale': 'Factor_Ven', 'Factor_Ventas': 'Factor_Ven',
    'Artículos por unidad de venta': 'Factor_Ven',
    # Stock
    'OnHand': 'Stock_SAP', 'Stock_SAP_Actual': 'Stock_SAP',
    'Stock disponible': 'Stock_SAP', 'En stock': 'Stock_SAP',
}
_RENAME_MOV = {
    # ItemCode — todas las variantes que SAP puede mostrar en español
    'ItemCode': 'ItemCode', 'Código de artículo': 'ItemCode',
    'Número de artículo': 'ItemCode', 'Artículo': 'ItemCode',
    'Cód. artículo': 'ItemCode', 'Codigo': 'ItemCode',
    'N.º artículo': 'ItemCode', 'Num artículo': 'ItemCode',
    'Cod. Artículo': 'ItemCode',
    # ItemName
    'ItemName': 'ItemName', 'Descripción del artículo': 'ItemName',
    'Nombre del artículo': 'ItemName', 'Nombre artículo': 'ItemName',
    'Nombre': 'ItemName', 'Descripcion': 'ItemName',
    # Fecha
    'Fecha': 'Fecha', 'DocDate': 'Fecha', 'Fecha de contabilización': 'Fecha',
    'Fecha doc.': 'Fecha', 'Fecha documento': 'Fecha',
    # Tipo transacción
    'Tipo': 'Tipo', 'Tipo_Doc': 'Tipo', 'Tipo de transacción': 'Tipo',
    # Referencia doc (BASE_REF)
    'N_Doc': 'N_Doc', 'BASE_REF': 'N_Doc', 'Ref. base': 'N_Doc',
    'DocNum': 'N_Doc', 'Número de documento': 'N_Doc',
    # Número de asiento (TransNum)
    'N_Asiento': 'N_Asiento', 'TransNum': 'N_Asiento',
    'Nº asiento': 'N_Asiento', 'Asiento': 'N_Asiento',
    # Cantidades
    'Entrada': 'Entrada', 'InQty': 'Entrada', 'Cantidad entrada': 'Entrada',
    'Salida': 'Salida', 'OutQty': 'Salida', 'Cantidad salida': 'Salida',
    'Neto': 'Neto',
    # Comentario
    'Comments': 'Comentario', 'Remark': 'Comentario', 'Comentario': 'Comentario',
    'Remarks': 'Comentario', 'Memo': 'Comentario',
    # Usuarios
    'Creado_Por': 'Creado_Por',
    'Modificado_Por': 'Modificado_Por', 'Modifico_Asiento': 'Modificado_Por',
    # Almacén
    'Almacen': 'Almacen', 'Almacén': 'Almacen', 'Warehouse': 'Almacen',
    'Depósito': 'Almacen', 'Bodega': 'Almacen',
}


def _renombrar(df: pd.DataFrame, mapa: dict) -> pd.DataFrame:
    df.columns = [c.strip() for c in df.columns]
    return df.rename(columns={k: v for k, v in mapa.items() if k in df.columns})


def _sfloat(val, default=0.0) -> float:
    try:
        f = float(val)
        return f if pd.notna(f) else default
    except Exception:
        return default


def _make_tree(parent_frame, cols, widths, height=14):
    col_ids = [f"c{i}" for i in range(len(cols))]
    tree = ttk.Treeview(parent_frame, columns=col_ids, show="headings", height=height)
    for cid, head, w in zip(col_ids, cols, widths):
        tree.heading(cid, text=head)
        tree.column(cid, width=w, anchor="center")
    sb_v = ttk.Scrollbar(parent_frame, orient="vertical",   command=tree.yview)
    sb_h = ttk.Scrollbar(parent_frame, orient="horizontal", command=tree.xview)
    tree.configure(yscroll=sb_v.set, xscroll=sb_h.set)
    sb_v.pack(side="right", fill="y")
    tree.pack(fill="both", expand=True)
    sb_h.pack(fill="x")
    return tree, col_ids


def _cargar_saldos_sig_db(item_tienda_pairs: list, fecha_desde: str, fecha_hasta: str):
    """
    Carga saldos desde SIG_DB filtrando por pares exactos (idproductos, idtienda).
    Devuelve (DataFrame, error_str | None).
    """
    import mysql.connector
    from config.settings import SIG_DB
    if not item_tienda_pairs:
        return pd.DataFrame(), "Sin pares artículo-tienda"
    ph = ','.join('(%s,%s)' for _ in item_tienda_pairs)
    args = [v for pair in item_tienda_pairs for v in pair]
    try:
        conn = mysql.connector.connect(**SIG_DB)
        cur  = conn.cursor(dictionary=True)
        cur.execute(f"""
            SELECT
                s.fproceso                              AS fecha,
                s.idproductos                           AS ItemCode,
                IFNULL(art.descrip1, s.idproductos)    AS descripcion,
                al.idtienda,
                al.descripcion                          AS nombre_almacen,
                s.idalmacen,
                IFNULL(ac.und_sap, '')                 AS und_sap,
                IFNULL(s.saldoinicial, 0)              AS saldoinicial,
                IFNULL(s.ingresos, 0)                  AS ingresos,
                IFNULL(s.salidas, 0)                   AS salidas
            FROM saldos s
            INNER JOIN almacenes al  ON al.idalmacen   = s.idalmacen
            LEFT  JOIN articulos art ON art.idarticulos = s.idproductos
            LEFT  JOIN almacart  ac
                   ON  ac.idalmacen   = s.idalmacen
                   AND ac.idarticulos = s.idproductos
                   AND ac.activo      = 1
            WHERE (s.idproductos, al.idtienda) IN ({ph})
              AND s.fproceso >= %s
              AND s.fproceso <= %s
            ORDER BY s.fproceso, s.idproductos, al.idtienda, s.idalmacen
        """, (*args, fecha_desde, fecha_hasta))
        rows = cur.fetchall()
        conn.close()
        return pd.DataFrame(rows), None
    except Exception as e:
        return pd.DataFrame(), str(e)


def _cargar_almacart(item_tienda_pairs: list):
    """Devuelve (df, error_str | None). Filtra por pares exactos (ItemCode, idtienda)."""
    import mysql.connector
    from config.settings import SIG_DB

    if not item_tienda_pairs:
        return pd.DataFrame(), "Sin pares artículo-tienda"
    ph = ','.join('(%s,%s)' for _ in item_tienda_pairs)
    args = [v for pair in item_tienda_pairs for v in pair]
    try:
        conn = mysql.connector.connect(**SIG_DB)
        cur  = conn.cursor(dictionary=True)
        cur.execute(f"""
            SELECT
                a.idarticulos           AS ItemCode,
                al.idtienda             AS idtienda,
                al.descripcion          AS nombre_almacen,
                IFNULL(a.und_sap, '')   AS und_sap,
                a.factor_sap,
                a.factorconversion      AS factor_sig,
                a.stockact,
                a.costoprom
            FROM almacart a
            INNER JOIN almacenes al ON al.idalmacen = a.idalmacen
            WHERE (a.idarticulos, al.idtienda) IN ({ph})
              AND a.activo = 1
            ORDER BY a.idarticulos, al.idtienda
        """, args)
        rows = cur.fetchall()
        conn.close()
        return pd.DataFrame(rows), None
    except Exception as e:
        return pd.DataFrame(), str(e)


# ────────────────────────────────────────────────────────────────────────────
# Editor de diccionario UOM
# ────────────────────────────────────────────────────────────────────────────

def _abrir_editor_uom(parent):
    """Ventana para ver/editar/guardar el diccionario SAP UOM → SIG UOM."""
    diccionario = _uom.cargar()

    ed = tb.Toplevel(parent)
    ed.title("Diccionario UOM — SAP → SIG")
    ed.geometry("640x520")
    set_window_icon(ed)

    tb.Label(ed,
             text="Mapeo de unidades de medida: cuando SAP muestra un nombre completo "
                  "(ej. 'KILOGRAMO') y SIG usa la abreviatura (ej. 'KG'), "
                  "defina aquí la equivalencia. La comparación usará este diccionario.",
             font=("Segoe UI", 8, "italic"), bootstyle="secondary",
             wraplength=620).pack(anchor="w", padx=12, pady=(8, 4))

    # ── Treeview del diccionario ──────────────────────────────────────────
    frame_t = tb.Frame(ed)
    frame_t.pack(fill="both", expand=True, padx=12, pady=(0, 4))
    tree_d = ttk.Treeview(frame_t, columns=("sap", "sig"), show="headings", height=14)
    tree_d.heading("sap", text="UOM en SAP (clave)")
    tree_d.heading("sig", text="Equivalente en SIG")
    tree_d.column("sap", width=260, anchor="w")
    tree_d.column("sig", width=260, anchor="w")
    sb = ttk.Scrollbar(frame_t, orient="vertical", command=tree_d.yview)
    tree_d.configure(yscroll=sb.set)
    sb.pack(side="right", fill="y")
    tree_d.pack(fill="both", expand=True)

    def _poblar():
        for iid in tree_d.get_children():
            tree_d.delete(iid)
        for sap_k, sig_v in sorted(diccionario.items()):
            tree_d.insert("", "end", values=(sap_k, sig_v))

    _poblar()

    # ── Fila de entrada ───────────────────────────────────────────────────
    row_add = tb.Frame(ed)
    row_add.pack(fill="x", padx=12, pady=(0, 4))
    tb.Label(row_add, text="SAP:").pack(side="left", padx=(0, 4))
    ent_sap = tb.Entry(row_add, width=22)
    ent_sap.pack(side="left", padx=(0, 8))
    tb.Label(row_add, text="SIG:").pack(side="left", padx=(0, 4))
    ent_sig = tb.Entry(row_add, width=22)
    ent_sig.pack(side="left", padx=(0, 8))

    def _agregar():
        k = ent_sap.get().strip().upper()
        v = ent_sig.get().strip()
        if not k or not v:
            messagebox.showwarning("Incompleto", "Complete ambos campos.", parent=ed)
            return
        diccionario[k] = v
        _poblar()
        ent_sap.delete(0, "end")
        ent_sig.delete(0, "end")

    def _eliminar():
        sel = tree_d.selection()
        if not sel:
            return
        for iid in sel:
            k = tree_d.item(iid, "values")[0]
            diccionario.pop(k, None)
        _poblar()

    def _editar(event):
        sel = tree_d.selection()
        if not sel:
            return
        vals = tree_d.item(sel[0], "values")
        ent_sap.delete(0, "end")
        ent_sap.insert(0, vals[0])
        ent_sig.delete(0, "end")
        ent_sig.insert(0, vals[1])

    tree_d.bind("<Double-1>", _editar)

    tb.Button(row_add, text="+ Agregar / Actualizar",
              bootstyle="success-outline", command=_agregar).pack(side="left", padx=(0, 6))
    tb.Button(row_add, text="✕ Eliminar seleccionado",
              bootstyle="danger-outline", command=_eliminar).pack(side="left")

    # ── Fila SAP UOM query + guardar ─────────────────────────────────────
    row_bot = tb.Frame(ed)
    row_bot.pack(fill="x", padx=12, pady=(0, 10))

    def _gen_uom():
        pyperclip.copy(construir_query_uom_sap())
        messagebox.showinfo("Query copiado",
                            "Query OUOM copiado.\n"
                            "Ejecútelo en SAP para ver los códigos y nombres de UOM.\n"
                            "Use esa información para completar el diccionario.",
                            parent=ed)

    def _guardar():
        _uom.guardar(diccionario)
        messagebox.showinfo(
            "Guardado",
            "Diccionario guardado en la base de datos SIG.\n"
            "Los cambios serán visibles en todos los equipos de inmediato.",
            parent=ed)
        ed.destroy()

    tb.Button(row_bot, text="📋 Query SAP — OUOM",
              bootstyle="info-outline", command=_gen_uom).pack(side="left", padx=(0, 8))
    tb.Button(row_bot, text="💾 Guardar y cerrar",
              bootstyle="success", command=_guardar).pack(side="left")
    tb.Button(row_bot, text="Cerrar sin guardar",
              bootstyle="secondary-outline",
              command=ed.destroy).pack(side="right")


# ────────────────────────────────────────────────────────────────────────────
# Panel principal
# ────────────────────────────────────────────────────────────────────────────

def abrir_panel_diagnostico(parent, df_ajustes: pd.DataFrame, razon_social: str = ""):
    items = sorted(df_ajustes['ItemCode'].astype(str).unique().tolist())

    # Pares exactos (ItemCode, tienda) y (ItemCode, WhsCode) — sin cruce entre tiendas
    _seen: set = set()
    item_tienda_pairs: list[tuple[str, str]] = []
    item_whs_pairs:    list[tuple[str, str]] = []
    for _, _r in df_ajustes.drop_duplicates(['ItemCode', 'ID_SIG']).iterrows():
        _item = str(_r['ItemCode'])
        try:
            _tid = str(int(float(str(_r['ID_SIG']))))
        except Exception:
            continue
        _whs = str(_r.get('WhsCode', ''))
        if (_item, _tid) not in _seen and _tid:
            _seen.add((_item, _tid))
            item_tienda_pairs.append((_item, _tid))
            if _whs:
                item_whs_pairs.append((_item, _whs))

    tiendas_ids = sorted({int(t) for _, t in item_tienda_pairs})

    if not items:
        messagebox.showinfo("Sin artículos", "No hay artículos para diagnosticar.")
        return

    win = tb.Toplevel(parent)
    win.title("Diagnóstico de Artículos — Análisis de Causa Raíz")
    win.geometry("1320x780")
    set_window_icon(win)

    hdr = tb.Frame(win, padding=(12, 6, 12, 2))
    hdr.pack(fill="x")
    tb.Label(hdr, text=f"Artículos ({len(items)}):",
             font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 6))
    tb.Label(hdr, text="  ·  ".join(items),
             font=("Segoe UI", 9), bootstyle="secondary").pack(side="left")
    btn_export = tb.Button(hdr, text="📥 Exportar Excel",
                           bootstyle="success-outline", width=18,
                           state="disabled")
    btn_export.pack(side="right", padx=(0, 4))
    ent_razon = tb.Entry(hdr, width=24, font=("Segoe UI", 9))
    ent_razon.pack(side="right", padx=(0, 4))
    if razon_social:
        ent_razon.insert(0, razon_social)
    tb.Label(hdr, text="Razón Social:",
             font=("Segoe UI", 9)).pack(side="right", padx=(8, 2))

    def _actualizar_btn_export():
        """Habilita exportar solo cuando los 3 pasos tienen datos cargados."""
        paso1_ok = _df_sap_cache[0] is not None
        paso2_ok = _df_mov_cache[0] is not None
        paso3_ok = not _sig3[0].empty and _sap3[0] is not None
        btn_export.config(
            state="normal" if (paso1_ok and paso2_ok and paso3_ok) else "disabled"
        )

    # Barra de error SIG — Text seleccionable, visible solo cuando hay mensaje
    txt_sig_err = tk.Text(win, height=2, bg="#3D0000", fg="#FF6B6B",
                          font=("Segoe UI", 8), wrap="word",
                          relief="flat", bd=4, cursor="arrow")
    txt_sig_err.bind("<Key>", lambda e: "break")   # seleccionable, no editable

    def _set_err(msg: str):
        txt_sig_err.config(state="normal")
        txt_sig_err.delete("1.0", "end")
        txt_sig_err.insert("1.0", msg)
        txt_sig_err.config(state="disabled")
        txt_sig_err.pack(fill="x", padx=10, pady=(0, 2))

    def _clear_err():
        txt_sig_err.pack_forget()

    nb = tb.Notebook(win, bootstyle="info")
    nb.pack(fill="both", expand=True, padx=10, pady=(4, 10))

    # ═══════════════════════════════════════════════════════════════════════
    # PASO 1 — Configuración SIG vs SAP
    # ═══════════════════════════════════════════════════════════════════════
    tab1 = tb.Frame(nb, padding=8)
    nb.add(tab1, text="  Paso 1 — Configuración UOM / Factor  ")

    tb.Label(tab1,
             text="Compara la configuración de unidad/factor en SIG (almacart, por tienda) "
                  "contra SAP (OITM). Una discrepancia explica diferencias sistemáticas de stock.",
             font=("Segoe UI", 8, "italic"), bootstyle="secondary",
             wraplength=1260).pack(anchor="w", pady=(0, 4))

    # ── Encabezado SIG ────────────────────────────────────────────────────
    row_sig_hdr = tb.Frame(tab1)
    row_sig_hdr.pack(fill="x", pady=(0, 2))
    lbl_sig_status = tb.Label(row_sig_hdr,
                               text="SIG — almacart  (cargando...)",
                               font=("Segoe UI", 9, "bold"))
    lbl_sig_status.pack(side="left")
    btn_reload_sig = tb.Button(row_sig_hdr, text="↺ Recargar SIG",
                                bootstyle="secondary-outline", width=16)
    btn_reload_sig.pack(side="left", padx=(10, 0))

    frame_sig = tb.Frame(tab1)
    frame_sig.pack(fill="x", pady=(0, 6))
    sig_cols   = ["ItemCode", "Tienda", "Almacén SIG", "UOM SAP (SIG)", "Factor SAP",
                  "Factor SIG", "Stock (almacart)", "Costo Prom."]
    sig_widths = [110, 60, 140, 120, 100, 100, 130, 100]
    tree_sig, _ = _make_tree(frame_sig, sig_cols, sig_widths, height=5)
    tree_sig.tag_configure('sin_conf', background='#3D1A00', foreground='#FFB347')
    tree_sig.tag_configure('normal',   background='')

    # Variables mutables para la comparación
    _sig_by_item: dict     = {}               # {ItemCode: {'und_sap': set, 'factor_sap': set}}
    _df_sig_store: list    = [pd.DataFrame()] # [DataFrame completo de SIG]
    _uom_mismatch_items: list = [set()]       # ItemCodes con UOM/Factor diferente en Paso 1

    def _poblar_sig(df_sig: pd.DataFrame):
        _sig_by_item.clear()
        _df_sig_store[0] = df_sig
        for iid in tree_sig.get_children():
            tree_sig.delete(iid)
        if df_sig.empty:
            tree_sig.insert("", "end", values=("Sin datos",) + ("—",) * 6)
            return
        for _, row in df_sig.iterrows():
            item    = str(row['ItemCode'])
            tienda  = str(row.get('idtienda', '—'))
            almacen = str(row.get('nombre_almacen', '—'))
            und     = str(row.get('und_sap', '') or '').strip()
            f_sap   = _sfloat(row.get('factor_sap'))
            f_sig   = _sfloat(row.get('factor_sig', 1.0))
            stk     = _sfloat(row.get('stockact'))
            costo   = _sfloat(row.get('costoprom'))
            tag     = 'sin_conf' if not und else 'normal'
            tree_sig.insert("", "end", tags=(tag,), values=(
                item, tienda, almacen,
                und or "⚠ Sin configurar",
                f"{f_sap:.4f}" if f_sap else "—",
                f"{f_sig:.4f}",
                f"{stk:.4f}",
                f"{costo:.4f}",
            ))
            if item not in _sig_by_item:
                _sig_by_item[item] = {'und_sap': set(), 'factor_sap': set()}
            if und:
                _sig_by_item[item]['und_sap'].add(und)
            if f_sap:
                _sig_by_item[item]['factor_sap'].add(round(f_sap, 4))

    def _cargar_sig():
        if not tiendas_ids:
            _set_err(
                "⚠  No se encontraron IDs de tienda en el resultado. "
                f"ID_SIG únicos: {list(df_ajustes['ID_SIG'].dropna().unique())[:10]}"
            )
            lbl_sig_status.config(text="SIG — almacart  (sin tiendas)", bootstyle="warning")
            return

        df_sig, err = _cargar_almacart(item_tienda_pairs)
        n = len(df_sig)

        if err:
            _set_err(f"⚠  {err}  |  items={items[:3]}  tiendas={tiendas_ids[:5]}")
            if n == 0:
                lbl_sig_status.config(text="SIG — almacart  (error de carga)", bootstyle="danger")
                _poblar_sig(pd.DataFrame())
                return
            # Datos parciales con advertencia (ej: tiendas sin und_sap)
            lbl_sig_status.config(
                text=f"SIG — almacart  ({n} filas, con advertencias)",
                bootstyle="warning")
        else:
            _clear_err()
            if n == 0:
                _set_err(
                    f"⚠  0 filas devueltas. "
                    f"items={items[:3]}  tiendas={tiendas_ids[:5]}"
                )
                lbl_sig_status.config(text="SIG — almacart  (sin datos)", bootstyle="warning")
            else:
                lbl_sig_status.config(
                    text=f"SIG — almacart  ({n} filas cargadas ✓)",
                    bootstyle="success")
        _poblar_sig(df_sig)

    btn_reload_sig.config(command=_cargar_sig)

    # ── SAP — OITM ────────────────────────────────────────────────────────
    row_sap_hdr = tb.Frame(tab1)
    row_sap_hdr.pack(fill="x", pady=(2, 2))
    tb.Label(row_sap_hdr, text="SAP — OITM (pegar resultado del query)",
             font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 12))
    btn1_gen     = tb.Button(row_sap_hdr, text="📋 Generar Query SAP",
                             bootstyle="info-outline", width=22)
    btn1_paste   = tb.Button(row_sap_hdr, text="📋 Pegar Resultado",
                             bootstyle="secondary", width=18)
    btn1_limpiar = tb.Button(row_sap_hdr, text="✕ Limpiar",
                             bootstyle="danger-outline", width=10)
    btn1_gen.pack(side="left", padx=(0, 4))
    btn1_paste.pack(side="left", padx=(0, 4))
    btn1_limpiar.pack(side="left", padx=(0, 16))
    tb.Button(row_sap_hdr, text="⚙ Diccionario UOM",
              bootstyle="warning-outline", width=18,
              command=lambda: _abrir_editor_uom(win)).pack(side="left")

    frame_sap = tb.Frame(tab1)
    frame_sap.pack(fill="x", pady=(2, 4))
    sap_cols   = ["ItemCode", "Descripción SAP", "UOM Inventario",
                  "UOM Compras", "Factor Compras", "UOM Ventas", "Factor Ventas", "Stock SAP"]
    sap_widths = [110, 210, 120, 110, 110, 110, 110, 110]
    tree_sap, _ = _make_tree(frame_sap, sap_cols, sap_widths, height=4)

    # ── Reporte de diferencias ────────────────────────────────────────────
    rep_hdr = tb.Frame(tab1)
    rep_hdr.pack(fill="x", pady=(2, 2))
    tb.Label(rep_hdr, text="Reporte de diferencias",
             font=("Segoe UI", 9, "bold")).pack(side="left")
    btn1_reproc = tb.Button(rep_hdr, text="↺ Reprocesar",
                            bootstyle="secondary-outline", width=14)
    btn1_reproc.pack(side="left", padx=(10, 0))

    frame_rep = tb.Frame(tab1)
    frame_rep.pack(fill="both", expand=True, pady=(0, 0))
    rep_cols   = ["ItemCode", "Nombre SAP", "Tienda", "Almacén", "UOM Inv SAP",
                  "Factor Compra SAP", "UOM SIG", "Factor SIG", "Descripción"]
    rep_widths = [110, 200, 60, 200, 130, 100, 100, 100, 280]
    tree_rep, _ = _make_tree(frame_rep, rep_cols, rep_widths, height=8)
    tree_rep.tag_configure('diff',    background='#3D1A00', foreground='#FFB347')
    tree_rep.tag_configure('sin_cfg', background='#2E0A0A', foreground='#FF8080')
    tree_rep.tag_configure('ok',      background='#0A2E0A', foreground='#6BCB77')

    _df_sap_cache: list = [None]

    def _analizar(df_sap: pd.DataFrame):
        # Cargar diccionario UOM actualizado (puede haber cambiado desde que se abrió)
        uom_dic = _uom.cargar()

        # Ordenar SAP por ItemCode
        df_sap = df_sap.sort_values('ItemCode', key=lambda s: s.astype(str),
                                    ignore_index=True)

        # Poblar tree_sap (datos crudos SAP)
        for iid in tree_sap.get_children():
            tree_sap.delete(iid)
        for _, row in df_sap.iterrows():
            tree_sap.insert("", "end", values=(
                str(row.get('ItemCode', '')),
                str(row.get('ItemName', '—')),
                str(row.get('UOM_Inv',  '—')).strip(),
                str(row.get('UOM_Com',  '—')).strip(),
                f"{_sfloat(row.get('Factor_Com', 1.0)) or 1.0:.4f}",
                str(row.get('UOM_Ven',  '—')).strip(),
                f"{_sfloat(row.get('Factor_Ven', 1.0)) or 1.0:.4f}",
                f"{_sfloat(row.get('Stock_SAP', 0.0)):.4f}",
            ))

        # Generar reporte por (ItemCode, Tienda, Almacén)
        for iid in tree_rep.get_children():
            tree_rep.delete(iid)

        df_sig = _df_sig_store[0]
        sap_idx = {str(r['ItemCode']): r for _, r in df_sap.iterrows()}

        # Items en SAP pero sin ninguna fila SIG
        for item in sorted(sap_idx):
            if df_sig.empty or item not in df_sig['ItemCode'].astype(str).values:
                tree_rep.insert("", "end", tags=('sin_cfg',), values=(
                    item, str(sap_idx[item].get('ItemName', '—')),
                    "—", "—", "—", "—", "—", "—",
                    "Sin datos SIG — recargar SIG",
                ))

        if df_sig.empty:
            return

        _uom_mismatch_items[0] = set()
        # Comparar por cada fila SIG
        for _, sig_row in df_sig.sort_values('ItemCode', key=lambda s: s.astype(str)).iterrows():
            item    = str(sig_row['ItemCode'])
            tienda  = str(sig_row.get('idtienda', '—'))
            almacen = str(sig_row.get('nombre_almacen', '—'))
            und_sig = str(sig_row.get('und_sap', '') or '').strip()
            f_sig   = _sfloat(sig_row.get('factor_sap'))

            sap_row = sap_idx.get(item)
            if sap_row is None:
                tree_rep.insert("", "end", tags=('sin_cfg',), values=(
                    item, "—", tienda, almacen, "—", "—",
                    und_sig or "—", f"{f_sig:.4f}" if f_sig else "—",
                    "Item no encontrado en resultado SAP pegado",
                ))
                continue

            # UOM inventario SAP (comparación de unidad)
            uom_inv_sap      = str(sap_row.get('UOM_Inv', '—')).strip()
            uom_inv_sap_norm = _uom.normalizar(uom_inv_sap, uom_dic)
            # Factor de compras SAP: NumInBuy (comparación de factor)
            f_com_sap = _sfloat(sap_row.get('Factor_Com', 1.0)) or 1.0

            difs = []
            tag  = 'ok'

            # Sin configuración SIG
            if not und_sig and not f_sig:
                difs.append("Sin und_sap ni factor_sap configurados en SIG")
                tag = 'sin_cfg'
            else:
                # Comparar UOM: und_sap (SIG) vs InvntryUom (SAP) normalizado
                if not und_sig:
                    difs.append("und_sap vacío en SIG")
                    tag = 'sin_cfg'
                elif uom_inv_sap not in ('—', '') and und_sig != uom_inv_sap_norm:
                    difs.append(
                        f"UOM inv SIG='{und_sig}' ≠ SAP='{uom_inv_sap}'"
                        + (f" (norm→'{uom_inv_sap_norm}')" if uom_inv_sap_norm != uom_inv_sap else "")
                    )
                    tag = 'diff'

                # Comparar factor: factor_sap (SIG) vs NumInBuy (SAP)
                if not f_sig:
                    difs.append("factor_sap sin configurar en SIG")
                    if tag == 'ok':
                        tag = 'sin_cfg'
                elif abs(f_sig - f_com_sap) > 0.001:
                    difs.append(f"Factor SIG={f_sig:.4f} ≠ NumInBuy SAP={f_com_sap:.4f}")
                    if tag == 'ok':
                        tag = 'diff'

            if tag in ('diff', 'sin_cfg'):
                _uom_mismatch_items[0].add(item)
            tree_rep.insert("", "end", tags=(tag,), values=(
                item, str(sap_row.get('ItemName', '—')),
                tienda, almacen,
                uom_inv_sap,
                f"{f_com_sap:.4f}",
                und_sig or "—",
                f"{f_sig:.4f}" if f_sig else "—",
                " | ".join(difs) if difs else "✓ Configuración correcta",
            ))

    def _gen_config():
        pyperclip.copy(construir_query_config(items))
        messagebox.showinfo("Query copiado",
                            "Query OITM copiado.\n"
                            "Ejecútelo en SAP y pegue con '📋 Pegar Resultado'.",
                            parent=win)

    def _paste_config():
        try:
            df = _renombrar(pd.read_clipboard(sep='\t'), _RENAME_OITM)
            _df_sap_cache[0] = df
            _analizar(df)
            _actualizar_btn_export()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo procesar:\n{e}", parent=win)

    def _limpiar_sap():
        _df_sap_cache[0] = None
        _uom_mismatch_items[0] = set()
        for t in (tree_sap, tree_rep):
            for iid in t.get_children():
                t.delete(iid)
        _actualizar_btn_export()

    def _reprocesar():
        if _df_sap_cache[0] is None:
            messagebox.showinfo("Sin datos SAP",
                                "Primero pegue el resultado SAP.", parent=win)
            return
        _analizar(_df_sap_cache[0])

    btn1_gen.config(command=_gen_config)
    btn1_paste.config(command=_paste_config)
    btn1_limpiar.config(command=_limpiar_sap)
    btn1_reproc.config(command=_reprocesar)

    # Carga inicial de SIG
    win.after(100, _cargar_sig)

    # ═══════════════════════════════════════════════════════════════════════
    # PASO 2 — Devoluciones / Anulaciones desde el 21/06
    # ═══════════════════════════════════════════════════════════════════════
    tab2 = tb.Frame(nb, padding=8)
    nb.add(tab2, text="  Paso 2 — Devoluciones / Anulaciones  ")

    tb.Label(tab2,
             text="Busca devoluciones (TransType 16), devoluciones a proveedor (21), "
                  "notas de crédito (14) y salidas manuales (60) desde la fecha indicada. "
                  "El 20/06/2026 fue la última corrección de salidas.",
             font=("Segoe UI", 8, "italic"), bootstyle="secondary",
             wraplength=1260).pack(anchor="w", pady=(0, 6))

    _var_filtro_manager = tk.BooleanVar(value=False)

    row2 = tb.Frame(tab2)
    row2.pack(fill="x", pady=(0, 6))
    tb.Label(row2, text="Desde:", font=("Segoe UI", 9)).pack(side="left", padx=(0, 4))
    ent2_desde = CampoFecha(row2, initialdate=date(2026, 6, 21))
    ent2_desde.pack(side="left", padx=(0, 14))
    btn2_gen   = tb.Button(row2, text="📋 Generar Query SAP (OINM)",
                           bootstyle="info-outline", width=28)
    btn2_paste = tb.Button(row2, text="📋 Pegar Resultado",
                           bootstyle="secondary", width=18)
    btn2_limpiar = tb.Button(row2, text="✕ Limpiar",
                             bootstyle="danger-outline", width=10)
    btn2_gen.pack(side="left", padx=(0, 4))
    btn2_paste.pack(side="left", padx=(0, 4))
    btn2_limpiar.pack(side="left")
    tb.Checkbutton(
        row2, text="Ocultar registros automáticos (manager sin modificar)",
        variable=_var_filtro_manager,
        command=lambda: _analizar_mov(_df_mov_cache[0]) if _df_mov_cache[0] is not None else None,
        bootstyle="warning-round-toggle",
    ).pack(side="left", padx=(16, 0))

    frame2 = tb.Frame(tab2)
    frame2.pack(fill="both", expand=True, pady=(4, 0))
    tree2, _ = _make_tree(frame2,
        ["ItemCode", "Descripción", "Fecha", "Tipo", "Ref. Doc", "N° Asiento",
         "Entrada", "Salida", "Neto", "Almacén", "Creado Por", "Modificado Por", "Comentario"],
        [110, 200, 100, 160, 90, 90, 100, 100, 100, 80, 100, 110, 260])
    tree2.tag_configure('dev',    background='#2E1A00', foreground='#FFB347')
    tree2.tag_configure('neg',    background='#2E0A0A', foreground='#FF8080')
    tree2.tag_configure('manual', background='#1A002E', foreground='#B388FF')
    tree2.tag_configure('normal', background='')

    find2_frame = tk.Frame(tab2, bg="#1a1a2e", bd=1, relief="solid")
    find2_frame.pack(fill="x", pady=(6, 0))
    lbl2_find = tk.Label(find2_frame, text="— Sin análisis aún —",
                         bg="#1a1a2e", fg="#555555",
                         font=("Segoe UI", 8, "italic"), anchor="w",
                         wraplength=1100, justify="left")
    lbl2_find.pack(side="left", fill="x", expand=True, padx=10, pady=6)

    _df_mov_cache: list = [None]

    def _analizar_mov(df: pd.DataFrame):
        tiene_comentario = 'Comentario' in df.columns
        n_filtradas = 0
        if _var_filtro_manager.get():
            def _es_auto(r):
                c = str(r.get('Creado_Por', '') or '').strip().lower()
                m = str(r.get('Modificado_Por', '') or '').strip()
                return c == 'manager' and m in ('', 'nan', '-', '—')
            mask = df.apply(_es_auto, axis=1)
            n_filtradas = int(mask.sum())
            df = df[~mask].copy()

        for iid in tree2.get_children():
            tree2.delete(iid)
        n_dev = n_neg = n_manual = 0
        resumen: dict = {}

        for _, row in df.iterrows():
            tipo   = str(row.get('Tipo', ''))
            neto   = _sfloat(row.get('Neto', 0.0))
            item   = str(row.get('ItemCode', '—'))
            desc     = str(row.get('ItemName', '') or '').strip() or '—'
            creado   = str(row.get('Creado_Por',    '') or '').strip() or '—'
            modif    = str(row.get('Modificado_Por', '') or '').strip() or '—'
            coment   = str(row.get('Comentario', '') or '') if tiene_comentario else '—'
            es_dev    = any(k in tipo for k in ('Dev', 'NC', 'Devoluci', 'Crédito'))
            es_manual = 'Manual' in tipo
            es_neg    = neto < 0
            if es_dev:    n_dev    += 1
            if es_manual: n_manual += 1
            if es_neg:    n_neg    += 1
            tag = ('dev' if es_dev else 'manual' if es_manual
                   else 'neg' if es_neg else 'normal')
            if item not in resumen:
                resumen[item] = []
            if es_dev or es_manual or es_neg:
                resumen[item].append(tipo)
            tree2.insert("", "end", tags=(tag,), values=(
                item, desc, row.get('Fecha', '—'), tipo,
                row.get('N_Doc', '—'), row.get('N_Asiento', '—'),
                f"{_sfloat(row.get('Entrada')):.4f}",
                f"{_sfloat(row.get('Salida')):.4f}",
                f"{neto:.4f}", row.get('Almacen', '—'),
                creado, modif, coment,
            ))

        partes, color = [], "#6BCB77"
        if n_filtradas:
            partes.append(
                f"ℹ  {n_filtradas} fila(s) ocultadas — creador 'manager' sin modificador (registros automáticos).")
        if n_dev:
            partes.append(f"⚠  {n_dev} devolución(es)/NC — ingresan stock en SAP sin registro en SIG.")
            color = "#FFB347"
        if n_manual:
            partes.append(f"⚠  {n_manual} salida(s) manual(es) — reducen stock SAP sin registro SIG.")
            color = "#B388FF"
        if n_neg and not n_dev and not n_manual:
            partes.append(f"⚠  {n_neg} movimiento(s) con neto negativo no clasificado.")
            color = "#FF8080"
        for it, tipos in resumen.items():
            if tipos:
                partes.append(f"   {it}: {', '.join(set(tipos))}")
        if not partes:
            partes.append("✓  Sin devoluciones, anulaciones ni salidas manuales en el período.")
        lbl2_find.config(text="\n".join(partes), fg=color)

    def _gen_negativos():
        fecha_sap = ent2_desde.get().replace('-', '')
        pyperclip.copy(construir_query_negativos(item_whs_pairs, fecha_sap))
        messagebox.showinfo("Query copiado",
                            f"Query OINM desde {ent2_desde.get()} copiado.",
                            parent=win)

    def _paste_negativos():
        try:
            df = _renombrar(pd.read_clipboard(sep='\t'), _RENAME_MOV)
            _df_mov_cache[0] = df
            _analizar_mov(df)
            _actualizar_btn_export()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo procesar:\n{e}", parent=win)

    def _limpiar_mov():
        _df_mov_cache[0] = None
        for iid in tree2.get_children():
            tree2.delete(iid)
        lbl2_find.config(text="— Datos limpiados —", fg="#555555")
        _actualizar_btn_export()

    btn2_gen.config(command=_gen_negativos)
    btn2_paste.config(command=_paste_negativos)
    btn2_limpiar.config(command=_limpiar_mov)

    # ═══════════════════════════════════════════════════════════════════════
    # PASO 3 — Saldos SIG vs SAP
    # ═══════════════════════════════════════════════════════════════════════
    tab3 = tb.Frame(nb, padding=8)
    nb.add(tab3, text="  Paso 3 — Saldos SIG vs SAP  ")

    tb.Label(tab3,
             text="Compara la tabla saldos del SIG (saldo inicial, ingresos, salidas por día) "
                  "contra los movimientos OINM de SAP. Identifica desde qué fecha divergen "
                  "las cantidades. El SIG es la fuente de referencia.",
             font=("Segoe UI", 8, "italic"), bootstyle="secondary",
             wraplength=1260).pack(anchor="w", pady=(0, 4))

    # ── Rango de fechas ───────────────────────────────────────────────────
    _hoy         = date.today()
    _primer_mes  = date(_hoy.year, _hoy.month, 1)

    row3_dates = tb.Frame(tab3)
    row3_dates.pack(fill="x", pady=(0, 4))
    tb.Label(row3_dates, text="Desde:", font=("Segoe UI", 9)).pack(side="left", padx=(0, 4))
    ent3_desde = CampoFecha(row3_dates, initialdate=_primer_mes)
    ent3_desde.pack(side="left", padx=(0, 12))
    tb.Label(row3_dates, text="Hasta:", font=("Segoe UI", 9)).pack(side="left", padx=(0, 4))
    ent3_hasta = CampoFecha(row3_dates, initialdate=_hoy)
    ent3_hasta.pack(side="left")

    # ── Controles ─────────────────────────────────────────────────────────
    row3_ctrl = tb.Frame(tab3)
    row3_ctrl.pack(fill="x", pady=(2, 4))

    lbl3_sig = tb.Label(row3_ctrl, text="SIG — saldos  (pendiente...)",
                        font=("Segoe UI", 9, "bold"))
    lbl3_sig.pack(side="left", padx=(0, 8))
    btn3_rsig    = tb.Button(row3_ctrl, text="↺ Recargar SIG",
                             bootstyle="secondary-outline", width=16)
    btn3_gen3    = tb.Button(row3_ctrl, text="📋 Generar Query SAP",
                             bootstyle="info-outline", width=22)
    btn3_paste3  = tb.Button(row3_ctrl, text="📋 Pegar SAP",
                             bootstyle="secondary", width=14)
    btn3_limp3   = tb.Button(row3_ctrl, text="✕ Limpiar",
                             bootstyle="danger-outline", width=10)
    btn3_rsig.pack(side="left",  padx=(0, 14))
    btn3_gen3.pack(side="left",  padx=(0, 4))
    btn3_paste3.pack(side="left", padx=(0, 4))
    btn3_limp3.pack(side="left")

    # ── Barra de estado SAP ───────────────────────────────────────────────
    lbl3_sap_status = tk.Label(tab3, text="SAP — sin datos pegados",
                                bg="#1a1a2e", fg="#555555",
                                font=("Segoe UI", 8, "italic"), anchor="w")
    lbl3_sap_status.pack(fill="x", padx=4, pady=(0, 2))

    # ── Tabla de comparación ──────────────────────────────────────────────
    frame3 = tb.Frame(tab3)
    frame3.pack(fill="both", expand=True, pady=(2, 0))

    _SAL_COLS   = ["Fecha", "ItemCode", "Descripción", "Almacén SIG", "Tienda",
                   "UOM SIG", "Saldo Ini SIG", "Ingresos SIG", "Salidas SIG",
                   "UOM SAP", "Saldo Ini SAP", "Ingresos SAP", "Salidas SAP",
                   "Observaciones"]
    _SAL_WIDTHS = [90, 90, 160, 130, 55, 65, 95, 90, 90, 65, 95, 90, 90, 220]
    tree3, _ = _make_tree(frame3, _SAL_COLS, _SAL_WIDTHS, height=20)
    tree3.tag_configure('diff',      background='#3D1A00', foreground='#FFB347')
    tree3.tag_configure('diff_soft', background='#2A2200', foreground='#D4C57E')
    tree3.tag_configure('sin_sap',   background='#1A1A3D', foreground='#7EB8F7')
    tree3.tag_configure('ok',        background='#0A2E0A', foreground='#6BCB77')

    # Estado interno
    _sig3: list = [pd.DataFrame()]
    _sap3: list = [None]
    _faltantes_sap: list = [[]]   # [(fecha, ItemCode, tienda), ...]

    _RENAME_SAP3 = {
        'ItemCode': 'ItemCode', 'Artículo': 'ItemCode',
        'Código de artículo': 'ItemCode', 'Cód. artículo': 'ItemCode',
        'N.º artículo': 'ItemCode',
        'Fecha': 'Fecha', 'DocDate': 'Fecha',
        'Almacen': 'Almacen', 'Almacén': 'Almacen', 'Warehouse': 'Almacen',
        'Ingresos_SAP': 'Ingresos_SAP', 'Ingresos': 'Ingresos_SAP',
        'Salidas_SAP':  'Salidas_SAP',  'Salidas':  'Salidas_SAP',
        'Stock_Previo': 'Stock_Previo', 'Stock previo': 'Stock_Previo',
        'UOM_Inv': 'UOM_Inv', 'InvntryUom': 'UOM_Inv',
        'UOM inventario': 'UOM_Inv', 'UM inventario': 'UOM_Inv',
    }

    # Mapping tienda (ID_SIG) → WhsCode SAP
    _tienda_whs: dict = {}
    for _, _r in df_ajustes.drop_duplicates(['ID_SIG', 'WhsCode']).iterrows():
        try:
            _tienda_whs[str(int(float(str(_r['ID_SIG']))))] = str(_r['WhsCode'])
        except Exception:
            pass

    def _merge_saldos(df: pd.DataFrame) -> pd.DataFrame:
        """Combina filas de distintos almacenes para el mismo (ItemCode, tienda, fecha)."""
        rows = []
        for (item, tienda, fecha), g in df.groupby(
                ['ItemCode', 'idtienda', 'fecha'], sort=True):
            almacenes  = list(g['nombre_almacen'].astype(str))
            uoms       = [str(u).strip() for u in g['und_sap']]
            uom_unicos = list(dict.fromkeys(u for u in uoms if u))
            base = {
                'fecha':         fecha,
                'ItemCode':      item,
                'descripcion':   str(g['descripcion'].iloc[0]),
                'idtienda':      tienda,
                'nombre_almacen': '/'.join(almacenes),
                '_mixed':        False,
            }
            if len(uom_unicos) <= 1:
                base['und_sap']      = uom_unicos[0] if uom_unicos else '—'
                base['saldoinicial'] = float(g['saldoinicial'].sum())
                base['ingresos']     = float(g['ingresos'].sum())
                base['salidas']      = float(g['salidas'].sum())
            else:
                base['_mixed']       = True
                base['und_sap']      = '/'.join(uom_unicos)
                base['saldoinicial'] = '/'.join(f"{float(v):.2f}" for v in g['saldoinicial'])
                base['ingresos']     = '/'.join(f"{float(v):.2f}" for v in g['ingresos'])
                base['salidas']      = '/'.join(f"{float(v):.2f}" for v in g['salidas'])
            rows.append(base)
        return pd.DataFrame(rows)

    def _poblar3():
        _faltantes_sap[0] = []
        for iid in tree3.get_children():
            tree3.delete(iid)
        df_sig = _sig3[0]
        df_sap = _sap3[0]
        if df_sig.empty:
            tree3.insert("", "end", values=("Sin datos SIG",) + ("—",) * 13)
            return

        # Normalizar columna Fecha del SAP a 'YYYY-MM-DD' (independiente del formato del clipboard)
        def _norm_fecha_col(df: pd.DataFrame) -> pd.DataFrame:
            if 'Fecha' not in df.columns:
                return df
            df = df.copy()
            df['Fecha'] = (pd.to_datetime(df['Fecha'], dayfirst=True, errors='coerce')
                           .dt.strftime('%Y-%m-%d'))
            return df

        # Índice SAP: (ItemCode, Warehouse, 'YYYY-MM-DD') → {saldo_ini, ingresos, salidas}
        sap_idx: dict = {}
        if df_sap is not None and not df_sap.empty:
            df_sap_n = _norm_fecha_col(df_sap)
            tiene_previo = 'Stock_Previo' in df_sap_n.columns
            for (item_s, whs_s), g in df_sap_n.groupby(['ItemCode', 'Almacen']):
                g = g.sort_values('Fecha')
                running = _sfloat(g['Stock_Previo'].iloc[0]) if tiene_previo else 0.0
                for _, r in g.iterrows():
                    fstr = str(r['Fecha'])[:10]
                    ing  = _sfloat(r.get('Ingresos_SAP', 0))
                    sal  = _sfloat(r.get('Salidas_SAP',  0))
                    sap_idx[(str(item_s), str(whs_s), fstr)] = {
                        'saldo_ini': running, 'ingresos': ing, 'salidas': sal,
                    }
                    running += ing - sal

        # UOM_Inv: viene en el resultado SAP pegado (CTE incluye OITM)
        # Descripción: desde caché Tab 1 si existe, sino ItemCode
        uom_sap_idx:  dict = {}
        desc_sap_idx: dict = {}
        if df_sap is not None and not df_sap.empty and 'UOM_Inv' in df_sap.columns:
            for item_k, g in df_sap.groupby('ItemCode'):
                uom_sap_idx[str(item_k)] = str(g['UOM_Inv'].iloc[0])
        if _df_sap_cache[0] is not None:
            for _, r in _df_sap_cache[0].iterrows():
                k = str(r.get('ItemCode', ''))
                if not uom_sap_idx.get(k):
                    uom_sap_idx[k] = str(r.get('UOM_Inv', '—'))
                desc_sap_idx[k] = str(r.get('ItemName', ''))

        df_merged = _merge_saldos(df_sig)

        for _, row in df_merged.iterrows():
            try:
                fecha_str = pd.Timestamp(row['fecha']).strftime('%Y-%m-%d')
            except Exception:
                fecha_str = str(row['fecha'])[:10]
            item      = str(row['ItemCode'])
            # Normalizar tienda: "024" → "24" para que coincida con _tienda_whs
            try:
                tienda = str(int(float(str(row['idtienda']))))
            except Exception:
                tienda = str(row['idtienda'])
            whs       = _tienda_whs.get(tienda, '')
            mixed     = bool(row.get('_mixed', False))
            si_sig    = row['saldoinicial']
            ing_sig   = row['ingresos']
            sal_sig   = row['salidas']

            sap_data = sap_idx.get((item, whs, fecha_str)) if whs else None
            uom_sap  = uom_sap_idx.get(item, '—')

            def _fs(v):
                return v if isinstance(v, str) else f"{float(v):.4f}"

            def _fv(sap_v, sig_v):
                if sap_v is None:
                    return '—'
                s = f"{sap_v:.4f}"
                if not mixed and not isinstance(sig_v, str):
                    if abs(sap_v - float(sig_v)) > 0.001:
                        return f"⚠ {s}"
                return s

            if sap_data:
                si_s  = _fv(sap_data['saldo_ini'], si_sig)
                ing_s = _fv(sap_data['ingresos'],  ing_sig)
                sal_s = _fv(sap_data['salidas'],    sal_sig)
                if '⚠' in ing_s:
                    tag = 'diff'
                elif '⚠' in si_s or '⚠' in sal_s:
                    tag = 'diff_soft'
                else:
                    tag = 'ok'

                obs_parts = []
                if '⚠' in si_s:
                    obs_parts.append("Dif. Saldo Inicial")
                if '⚠' in ing_s:
                    if not mixed and not isinstance(ing_sig, str):
                        obs_parts.append(
                            "Falta ingreso en SAP"
                            if sap_data['ingresos'] < float(ing_sig)
                            else "Ingreso extra en SAP"
                        )
                    else:
                        obs_parts.append("Dif. Ingresos")
                if '⚠' in sal_s:
                    if not mixed and not isinstance(sal_sig, str):
                        obs_parts.append(
                            "Falta salida en SAP"
                            if sap_data['salidas'] < float(sal_sig)
                            else "Salida extra en SAP"
                        )
                    else:
                        obs_parts.append("Dif. Salidas")
                obs = " | ".join(obs_parts) if obs_parts else "✓ Sin diferencias"
            else:
                si_s = ing_s = sal_s = '—'
                tag  = 'sin_sap'
                obs  = "Sin datos SAP"

            if 'Falta ingreso en SAP' in obs and item not in _uom_mismatch_items[0]:
                _faltantes_sap[0].append((fecha_str, item, tienda))
            desc = desc_sap_idx.get(item) or str(row.get('descripcion', item))
            tree3.insert("", "end", tags=(tag,), values=(
                fecha_str, item,
                desc,
                str(row['nombre_almacen']),
                tienda,
                str(row['und_sap']),
                _fs(si_sig), _fs(ing_sig), _fs(sal_sig),
                uom_sap, si_s, ing_s, sal_s,
                obs,
            ))

    def _cargar_sig3():
        fd = ent3_desde.get()
        fh = ent3_hasta.get()
        lbl3_sig.config(text="SIG — saldos  (cargando...)", bootstyle="secondary")
        df, err = _cargar_saldos_sig_db(item_tienda_pairs, fd, fh)
        if err:
            _set_err(f"⚠ saldos SIG: {err}")
            lbl3_sig.config(text="SIG — saldos  (error)", bootstyle="danger")
            _sig3[0] = pd.DataFrame()
        else:
            _clear_err()
            n = len(df)
            lbl3_sig.config(text=f"SIG — saldos  ({n} filas ✓)", bootstyle="success")
            _sig3[0] = df
        _poblar3()
        _actualizar_btn_export()
        _actualizar_tab4()

    def _gen_sap3():
        fd = ent3_desde.get().replace('-', '')
        fh = ent3_hasta.get().replace('-', '')
        pyperclip.copy(construir_query_saldos_sap(item_whs_pairs, fd, fh))
        messagebox.showinfo(
            "Query copiado",
            f"Query OINM {ent3_desde.get()} → {ent3_hasta.get()} copiado.\n"
            "Incluye stock previo al período (Stock_Previo).\n"
            "Ejecute en SAP y pegue con '📋 Pegar SAP'.",
            parent=win)

    def _paste_sap3():
        try:
            df = _renombrar(pd.read_clipboard(sep='\t'), _RENAME_SAP3)
            cols_req = {'ItemCode', 'Almacen', 'Ingresos_SAP', 'Salidas_SAP'}
            faltantes = cols_req - set(df.columns)
            if faltantes:
                messagebox.showwarning(
                    "Columnas faltantes",
                    f"El resultado pegado no tiene: {', '.join(sorted(faltantes))}\n\n"
                    "Verifique que:\n"
                    "1. Copió el resultado del query SAP generado aquí.\n"
                    "2. Los encabezados de columna son los del query (ItemCode, Almacen, etc.).",
                    parent=win)
                return
            _sap3[0] = df
            n = len(df)
            items_sap = df['ItemCode'].nunique()
            # Muestra un ejemplo de clave para que el usuario pueda verificar el formato
            ej_item = df['ItemCode'].iloc[0] if n > 0 else '—'
            ej_alm  = df['Almacen'].iloc[0]  if n > 0 else '—'
            ej_fech = str(df['Fecha'].iloc[0])[:10] if n > 0 else '—'
            lbl3_sap_status.config(
                text=f"SAP — {n} filas · {items_sap} art. · ej: {ej_item} | {ej_alm} | {ej_fech}",
                fg="#6BCB77")
            _poblar3()
            _actualizar_btn_export()
            _actualizar_tab4()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo procesar:\n{e}", parent=win)

    def _limpiar3():
        _sap3[0] = None
        _sig3[0] = pd.DataFrame()
        _faltantes_sap[0] = []
        for iid in tree3.get_children():
            tree3.delete(iid)
        lbl3_sig.config(text="SIG — saldos  (limpiado)", bootstyle="secondary")
        lbl3_sap_status.config(text="SAP — sin datos pegados", fg="#555555")
        _actualizar_btn_export()
        _actualizar_tab4()

    btn3_rsig.config(command=_cargar_sig3)
    btn3_gen3.config(command=_gen_sap3)
    btn3_paste3.config(command=_paste_sap3)
    btn3_limp3.config(command=_limpiar3)

    win.after(200, _cargar_sig3)

    # ═══════════════════════════════════════════════════════════════════════
    # PASO 4 — Ingresos SIG sin migrar a SAP
    # ═══════════════════════════════════════════════════════════════════════
    tab4 = tb.Frame(nb, padding=8)
    nb.add(tab4, text="  Paso 4 — Ingresos SIG sin migrar  ")

    tb.Label(tab4,
             text="Ingresos registrados en SIG (ingresosprod) para los casos donde se detectó "
                  "'Falta ingreso en SAP' en el Paso 3. Permite identificar qué documentos "
                  "de SIG aún no se han migrado al inventario SAP.",
             font=("Segoe UI", 8, "italic"), bootstyle="secondary",
             wraplength=1260).pack(anchor="w", pady=(0, 4))

    row4_ctrl = tb.Frame(tab4)
    row4_ctrl.pack(fill="x", pady=(0, 6))
    lbl4_status = tb.Label(row4_ctrl, text="Esperando análisis del Paso 3...",
                           font=("Segoe UI", 9), bootstyle="secondary")
    lbl4_status.pack(side="left", padx=(0, 12))
    btn4_cargar = tb.Button(row4_ctrl, text="🔍 Cargar ingresos SIG faltantes",
                            bootstyle="info-outline", width=32, state="disabled")
    btn4_cargar.pack(side="left")

    frame4 = tb.Frame(tab4)
    frame4.pack(fill="both", expand=True, pady=(4, 0))
    _TAB4_COLS = ["Fecha", "ItemCode", "Descripción", "Tienda",
                  "N° Ingreso", "N° Documento", "Cantidad", "Obs. SIG"]
    tree4, _ = _make_tree(frame4, _TAB4_COLS,
                          [90, 90, 200, 60, 90, 110, 90, 280], height=20)
    tree4.tag_configure('faltante', background='#3D1A00', foreground='#FFB347')

    def _actualizar_tab4():
        casos = list(set(_faltantes_sap[0]))
        if casos:
            lbl4_status.config(
                text=f"⚠ {len(casos)} caso(s) 'Falta ingreso en SAP' — pulsar para cargar detalle SIG",
                bootstyle="warning")
            btn4_cargar.config(state="normal")
        else:
            lbl4_status.config(
                text="✓ Sin casos 'Falta ingreso en SAP' detectados en Paso 3",
                bootstyle="success")
            btn4_cargar.config(state="disabled")
            for iid in tree4.get_children():
                tree4.delete(iid)

    def _cargar_ingresos_faltantes():
        import mysql.connector
        from config.settings import SIG_DB
        casos = list(set(_faltantes_sap[0]))
        if not casos:
            return
        ph   = ','.join('(%s,%s,%s)' for _ in casos)
        args = []
        for fecha, item_c, tid in casos:
            try:
                args.extend([fecha, int(item_c), int(tid)])
            except ValueError:
                args.extend([fecha, item_c, tid])
        try:
            conn = mysql.connector.connect(**SIG_DB)
            cur  = conn.cursor(dictionary=True)
            cur.execute(f"""
                SELECT
                    ip.fproceso                                         AS fecha,
                    ipd.idingprod,
                    ip.nrodocum                                         AS nrodocum,
                    CAST(ipd.idarticulos AS CHAR)                       AS ItemCode,
                    IFNULL(art.descrip1, CAST(ipd.idarticulos AS CHAR)) AS descripcion,
                    ip.idtienda,
                    ipd.cantidad,
                    IFNULL(ipd.obs, '')                                 AS obs
                FROM ingresosprod ip
                INNER JOIN ingproddetalle ipd ON ip.idingprod   = ipd.idingprod
                LEFT  JOIN articulos art      ON art.idarticulos = ipd.idarticulos
                WHERE (ip.fproceso, ipd.idarticulos, ip.idtienda) IN ({ph})
                ORDER BY ip.fproceso, ipd.idarticulos, ip.idtienda, ip.idingprod
            """, args)
            rows = cur.fetchall()
            conn.close()
            for iid in tree4.get_children():
                tree4.delete(iid)
            if not rows:
                lbl4_status.config(
                    text="Sin ingresos SIG encontrados para los casos detectados.",
                    bootstyle="warning")
                return
            for r in rows:
                tree4.insert("", "end", tags=('faltante',), values=(
                    str(r['fecha']),
                    str(r['ItemCode']),
                    str(r['descripcion']),
                    str(r['idtienda']),
                    str(r['idingprod']),
                    str(r['nrodocum']),
                    f"{float(r['cantidad']):.4f}",
                    str(r['obs']),
                ))
            lbl4_status.config(
                text=f"✓ {len(rows)} registro(s) SIG — documentos pendientes de migrar a SAP",
                bootstyle="success")
        except Exception as e:
            lbl4_status.config(text=f"Error SIG: {e}", bootstyle="danger")

    btn4_cargar.config(command=_cargar_ingresos_faltantes)

    # ═══════════════════════════════════════════════════════════════════════
    # Exportar a Excel — openpyxl con colores y valores numéricos limpios
    # ═══════════════════════════════════════════════════════════════════════
    def _exportar_excel():
        import re
        from datetime import datetime as _dt
        from tkinter import filedialog
        import openpyxl
        from openpyxl.styles import PatternFill, Font

        razon      = ent_razon.get().strip()
        razon_safe = re.sub(r'[^\w]', '', razon.replace(' ', ''))[:30]
        ts         = _dt.now().strftime('%H%M%d%m%Y')
        fname      = (f"diagnostico_{razon_safe}_{ts}.xlsx"
                      if razon_safe else f"diagnostico_{ts}.xlsx")

        path = filedialog.asksaveasfilename(
            parent=win,
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Guardar diagnóstico en Excel",
            initialfile=fname,
        )
        if not path:
            return

        # tag → (bg_hex, fg_hex) — mismos colores que los tag_configure
        _TAG_COLORS = {
            'diff':      ('3D1A00', 'FFB347'),
            'diff_soft': ('2A2200', 'D4C57E'),
            'sin_sap':   ('1A1A3D', '7EB8F7'),
            'ok':        ('0A2E0A', '6BCB77'),
            'sin_cfg':   ('2E0A0A', 'FF8080'),
            'dev':       ('2E1A00', 'FFB347'),
            'neg':       ('2E0A0A', 'FF8080'),
            'manual':    ('1A002E', 'B388FF'),
            'faltante':  ('3D1A00', 'FFB347'),
        }
        _NUM_COLS = {
            'Saldo Ini SIG', 'Ingresos SIG', 'Salidas SIG',
            'Saldo Ini SAP', 'Ingresos SAP', 'Salidas SAP',
            'Entrada', 'Salida', 'Neto', 'Cantidad',
            'Factor Compra SAP', 'Factor SIG',
        }

        def _clean(v: str) -> str:
            s = str(v).strip()
            return s[1:].strip() if s.startswith('⚠') else s

        def _cell_val(col_name: str, raw):
            s = _clean(str(raw))
            if col_name in _NUM_COLS:
                try:
                    return float(s)
                except (ValueError, TypeError):
                    pass
            return s if s not in ('—', '', 'None') else None

        _HDR_FILL = PatternFill('solid', fgColor='1E3A5F')
        _HDR_FONT = Font(bold=True, color='FFFFFF', size=9)

        def _write_sheet(wb, sheet_name, tree, col_names):
            ws = wb.create_sheet(sheet_name)
            for ci, name in enumerate(col_names, 1):
                c = ws.cell(row=1, column=ci, value=name)
                c.fill = _HDR_FILL
                c.font = _HDR_FONT
            for ri, iid in enumerate(tree.get_children(), 2):
                vals = list(tree.item(iid, 'values'))
                tags = tree.item(iid, 'tags')
                tag  = tags[0] if tags else ''
                bg, fg = _TAG_COLORS.get(tag, (None, None))
                fill = PatternFill('solid', fgColor=bg) if bg else None
                fnt  = Font(color=fg, size=9) if fg else Font(size=9)
                for ci, col_name in enumerate(col_names, 1):
                    raw = vals[ci - 1] if ci - 1 < len(vals) else ''
                    c   = ws.cell(row=ri, column=ci, value=_cell_val(col_name, raw))
                    if fill:
                        c.fill = fill
                    c.font = fnt
            for col in ws.columns:
                max_w = max((len(str(c.value or '')) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_w + 2, 45)

        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            _write_sheet(wb, 'Paso1_Config',       tree_rep, rep_cols)
            _write_sheet(wb, 'Paso2_Devoluciones', tree2, [
                "ItemCode", "Descripción", "Fecha", "Tipo", "Ref. Doc", "N° Asiento",
                "Entrada", "Salida", "Neto", "Almacén",
                "Creado Por", "Modificado Por", "Comentario",
            ])
            _write_sheet(wb, 'Paso3_Saldos',       tree3, _SAL_COLS)
            if tree4.get_children():
                _write_sheet(wb, 'Paso4_Ingresos_SIG', tree4, _TAB4_COLS)
            wb.save(path)
            messagebox.showinfo("Exportado",
                                f"Diagnóstico guardado:\n{path}", parent=win)
        except Exception as e:
            messagebox.showerror("Error al exportar", str(e), parent=win)

    btn_export.config(command=_exportar_excel)
