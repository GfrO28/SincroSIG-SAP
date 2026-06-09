# modules/ report_formulaciones

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
from decimal import Decimal, InvalidOperation
from contextlib import contextmanager
from copy import copy
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

from config.utils import obtener_tiendas_sig
from config.db import get_store_connection
from modules.preciart_service import articulo_valido_para_formulacion, obtener_info_articulo
from modules.formulaciones_service import (
    obtener_idformulacion_por_articulo,
    obtener_detalle_formulacion,
    obtener_detalles_por_tienda,
    obtener_idformulacion_batch,
    obtener_detalles_formulaciones_batch
)

EPS = Decimal("0.01")
AUDITORIA = True

def _to_decimal(v):
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError, ValueError):
        return None


@contextmanager
def conexion_tienda(idtienda: int):
    """
    Context manager seguro para abrir y cerrar conexiones a tienda.
    """
    conn = cur = None
    try:
        conn, cur = get_store_connection(idtienda)
        yield conn, cur
    finally:
        if cur:
            try:
                cur.close()
            except Exception:
                pass
        if conn:
            try:
                conn.close()
            except Exception:
                pass

def construir_bloque_excel(
    detalles_por_tienda,
    tiendas,
    base_por_tipo,
    tipo_por_tienda,
    descripcion_formula,
    EPS
):
    bloque_excel = []
    bloque_flags = []
    hay_diferencias = False

    # 🔹 TITLE
    bloque_excel.append([descripcion_formula])
    bloque_flags.append("TITLE")

    # 🔹 HEADER
    encabezado = ["Insumo", "Unidad", "Almacén Base"] + [str(t) for t in tiendas]
    bloque_excel.append(encabezado)
    bloque_flags.append("HEADER")

    # 🔹 INSUMOS
    insumos_union = sorted(
        {k for d in detalles_por_tienda.values() for k in d.keys()}
    )

    base_por_tienda_real = {
        t: base_por_tipo.get(tipo_por_tienda.get(t, 1), None)
        for t in tiendas
    }

    #print("BASES:", base_por_tienda_real)
    #print("DETALLES KEYS:", detalles_por_tienda.keys())
    for id_ins in insumos_union:

        nombre = ""
        unidad = ""
        almacen_base = ""

        # 🔹 obtener datos base
        for d in detalles_por_tienda.values():
            if id_ins in d:
                nombre = d[id_ins]["insumo"]
                unidad = d[id_ins]["unidad"]
                break

        for base_tda in base_por_tienda_real.values():
            base_det = detalles_por_tienda.get(int(base_tda), {})
            if id_ins in base_det:
                almacen_base = base_det[id_ins]["almacen"]
                break

        fila_excel = [nombre, unidad, almacen_base]
        fila_flags = []

        for tda in tiendas:

            dtda = detalles_por_tienda.get(int(tda), {})
            base_tda = base_por_tienda_real.get(tda)
            print(f"BASE REAL PARA {tda}: {base_tda}")
            print("DETALLES DISPONIBLES:", detalles_por_tienda.keys())
            base_det = detalles_por_tienda.get(int(base_tda), {})

            logs = []
            mostrar_detalle = False

            if id_ins in dtda and id_ins in base_det:
                bc = _to_decimal(base_det[id_ins]["cantidad"])
                tc = _to_decimal(dtda[id_ins]["cantidad"])

                if (
                    bc is None or tc is None or
                    (bc - tc).copy_abs() > EPS or
                    base_det[id_ins]["unidad"].strip().upper() != dtda[id_ins]["unidad"].strip().upper() or
                    base_det[id_ins]["almacen"].upper() != dtda[id_ins]["almacen"].upper()
                ):
                    hay_diferencias = True
                    mostrar_detalle = True

            elif id_ins in dtda or id_ins in base_det:
                hay_diferencias = True
                mostrar_detalle = True

            # 🔥 AUDITORÍA
            if AUDITORIA:
                logs = auditar_diferencia(id_ins, tda, base_tda, base_det, dtda)
                for l in logs:
                    print(f"[AUDIT] Art:{id_ins} → {l}")

            # 🔹 valor a mostrar
            if id_ins in dtda:
                dd = dtda[id_ins]
                val = (
                    f"{float(dd['cantidad']):.4f}"
                    if not mostrar_detalle
                    else f"{float(dd['cantidad']):.4f} | {dd['unidad']} | {dd['almacen']}"
                )
            else:
                val = "-"

            fila_excel.append(val)
            fila_flags.append(mostrar_detalle)

        bloque_excel.append(fila_excel)
        bloque_flags.append(fila_flags)

    return bloque_excel, bloque_flags, hay_diferencias

def auditar_diferencia(id_ins, tda, base_tda, base_det, dtda):
    logs = []

    base = base_det.get(id_ins)
    tda_val = dtda.get(id_ins)

    if base and tda_val:
        bc = _to_decimal(base["cantidad"])
        tc = _to_decimal(tda_val["cantidad"])

        if bc is None or tc is None:
            logs.append(f"T{tda} vs B{base_tda} → ERROR CONVERSIÓN")
        else:
            if (bc - tc).copy_abs() > EPS:
                logs.append(
                    f"T{tda} vs B{base_tda} → DIF CANTIDAD ({tc} vs {bc})"
                )

        if base["unidad"].strip().upper() != tda_val["unidad"].strip().upper():
            logs.append(
                f"T{tda} vs B{base_tda} → DIF UNIDAD ({tda_val['unidad']} vs {base['unidad']})"
            )

        if base["almacen"].upper() != tda_val["almacen"].upper():
            logs.append(
                f"T{tda} vs B{base_tda} → DIF ALMACÉN ({tda_val['almacen']} vs {base['almacen']})"
            )

    elif base and not tda_val:
        logs.append(f"T{tda} vs B{base_tda} → FALTA EN TIENDA")

    elif tda_val and base is None:
        logs.append(f"T{tda} vs B{base_tda} → BASE SIN FORMULA")
    elif tda_val and base == {}:
        logs.append(f"T{tda} vs B{base_tda} → BASE SIN DETALLE")

    return logs

def generar_reporte_formulaciones(
    tiendas: list[int],
    articulos: list[int],
    ruta_salida: str | None = None,
    progress_callback=None,
    solo_diferencias: bool = False
):

    base_por_tipo = {1: 3, 2: 34, 3: 41}
    tiendas_info = obtener_tiendas_sig()
    tipo_por_tienda = {int(t[0]): int(t[2]) for t in tiendas_info}

    bold = Font(bold=True)
    white_font = Font(color="FFFFFF")
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws_diff = wb.active
    ws_diff.title = "DIFERENCIAS"

    ws_ok = None
    ws_inc = None

    if not solo_diferencias:
        ws_ok = wb.create_sheet("FORMULACIONES")
        ws_inc = wb.create_sheet("FALTANTES EN TDA BASE")

    faltantes = []
    cache_formulaciones = {}

    # ==========================================
    # 🔥 PRECIART
    # ==========================================
    cache_preciart = {}
    conn_val, cur_val = get_store_connection(tiendas[0])

    total = len(articulos)

    # ==========================================
    # 🔥 BASES
    # ==========================================
    tipos_en_tiendas = {tipo_por_tienda.get(t, 1) for t in tiendas}
    bases_requeridas_global = {base_por_tipo[t] for t in tipos_en_tiendas if t in base_por_tipo}
    tiendas_cache = list(set(tiendas) | bases_requeridas_global)

    # ==========================================
    # 🔥 CACHE IDFORMULACIONES
    # ==========================================
    cache_idforms = {}

    def cargar_idforms_tienda(tda):
        try:
            conn, cur = get_store_connection(tda)
            data = obtener_idformulacion_batch(cur, tda, articulos)
            return tda, data
        except Exception as e:
            print(f"⚠️ Error precarga tienda {tda}: {e}")
            return tda, {}
        finally:
            try:
                cur.close()
                conn.close()
            except:
                pass

    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = [executor.submit(cargar_idforms_tienda, t) for t in tiendas_cache]

        for future in as_completed(futures):
            tda, data = future.result()
            for idart_b, idform in data.items():
                cache_idforms[(int(idart_b), int(tda))] = idform

    print("DEBUG cache_idforms:", len(cache_idforms))

    # ==========================================
    # 🔥 CACHE DETALLES (CORREGIDO)
    # ==========================================
    cache_detalles = {}

    def cargar_detalles_tienda(tda):
        conn, cur = get_store_connection(tda)

        data = {}

        for idart in articulos:
            idform = cache_idforms.get((idart, tda))

            if not idform:
                continue

            detalle, desc = obtener_detalle_formulacion(cur, idform)

            if detalle:
                data[idart] = {
                    "detalles": detalle,
                    "descripcion": desc
                }

        return tda, data

    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = [executor.submit(cargar_detalles_tienda, t) for t in tiendas_cache]

        for future in as_completed(futures):
            tda, data = future.result()

            for idart, info in data.items():
                cache_detalles[(int(tda), idform)] = {
                    "detalles": info["detalles"],
                    "descripcion": info["descripcion"]
                }

    print("DEBUG cache_detalles:", len(cache_detalles))

    # ==========================================
    # 🔥 LOOP PRINCIPAL
    # ==========================================
    for idx, idart in enumerate(articulos, start=1):

        if not articulo_valido_para_formulacion(cur_val, idart, cache_preciart):
            if not solo_diferencias:
                ws_inc.append([idart, None, "GLOBAL", "NO EXISTE O INACTIVO EN PRECIART"])
            continue

        if progress_callback:
            progress_callback((idx / total) * 100, f"Procesando artículo {idx}/{total}")

        print(f"\n🔎 ART {idart}")

        tipos_en_tiendas = {tipo_por_tienda.get(t, 1) for t in tiendas}
        bases_requeridas = {base_por_tipo[t] for t in tipos_en_tiendas if t in base_por_tipo}

        tiendas_a_consultar = set(tiendas) | bases_requeridas

        detalles_por_tienda, descripcion_formula, faltantes_tmp = obtener_detalles_por_tienda(
            idart,
            tiendas_a_consultar,
            base_por_tipo,
            tipo_por_tienda,
            cache_formulaciones,
            cache_idforms,
            cache_detalles
        )

        print("BASES REQUERIDAS:", bases_requeridas)
        print("DETALLES DISPONIBLES:", detalles_por_tienda.keys())

        if not detalles_por_tienda:
            continue

        # 🔥 VALIDAR BASE
        base_inexistente = False
        for base_id in bases_requeridas:
            if base_id not in detalles_por_tienda:
                print(f"[ERROR] BASE NO CARGADA {base_id}")
                base_inexistente = True
                break

        if base_inexistente:
            if not solo_diferencias:
                ws_inc.append([idart, None, "BASE ERROR", "NO EXISTE EN BASE"])
            continue

        if not solo_diferencias:
            faltantes.extend(faltantes_tmp)

        bloque_excel, bloque_flags, hay_diferencias = construir_bloque_excel(
            detalles_por_tienda,
            tiendas,
            base_por_tipo,
            tipo_por_tienda,
            descripcion_formula,
            EPS
        )

        if solo_diferencias and not hay_diferencias:
            continue

        ws_target = ws_diff if hay_diferencias else ws_ok

        for fila_idx, fila in enumerate(bloque_excel):
            ws_target.append(fila)
            row_idx = ws_target.max_row
            flag = bloque_flags[fila_idx]

            if flag == "TITLE":
                ws_target.cell(row=row_idx, column=1).font = bold

            elif flag == "HEADER":
                for col in range(1, len(fila) + 1):
                    c = ws_target.cell(row=row_idx, column=col)
                    c.font = bold
                    c.fill = header_fill
                    c.alignment = center

            else:
                for i, diff in enumerate(flag):
                    if diff:
                        c = ws_target.cell(row=row_idx, column=4 + i)
                        c.font = white_font
                        c.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        ws_target.append([])

    cur_val.close()
    conn_val.close()

    if not ruta_salida:
        ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        ruta_salida = f"SincroSIG/logs/formulaciones/reporte_formulaciones_{ts}.xlsx"

    os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
    wb.save(ruta_salida)

    return ruta_salida