# modules/formulaciones_export.py

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

BOLD        = Font(bold=True)
WHITE_FONT  = Font(color="FFFFFF", bold=True)

HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
BASE_FILL   = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
DIFF_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

CENTER = Alignment(horizontal="center", vertical="center")


def exportar_excel_dataset(
    dataset: dict,
    tiendas: list,
    solo_diferencias: bool = False,
    base_por_tipo: dict | None = None,
    ruta_salida: str | None = None
):
    """
    base_por_tipo = {1: 3, 2: 34, 3: 41}
    En modo solo_diferencias las columnas son:
        [tiendas base presentes en la selección] + [tiendas con al menos 1 diff]
    En modo completo: todas las tiendas seleccionadas.
    """

    # Conjunto de tiendas base (3, 34, 41) que estén en la selección
    bases_globales = set()
    if base_por_tipo:
        bases_globales = {int(b) for b in base_por_tipo.values()} & {int(t) for t in tiendas}

    wb = Workbook()

    ws_diff = wb.active
    ws_diff.title = "DIFERENCIAS"

    ws_ok = None
    if not solo_diferencias:
        ws_ok = wb.create_sheet("FORMULACIONES OK")

    for formula in dataset["articulos"]:

        if solo_diferencias and not formula["tiene_diferencias"]:
            continue

        # ==========================================
        # COLUMNAS POR FÓRMULA
        # ==========================================
        if solo_diferencias:
            # Tiendas que tienen al menos un insumo con diff
            tiendas_con_diff = set()
            for ins in formula["insumos"]:
                for tda, val in ins["tiendas"].items():
                    if val["diff"]:
                        tiendas_con_diff.add(int(tda))

            # Bases primero (ordenadas), luego diffs sin repetir bases
            columnas_tiendas = (
                sorted(bases_globales)
                + sorted(tiendas_con_diff - bases_globales)
            )
        else:
            columnas_tiendas = [int(t) for t in tiendas]

        # ==========================================
        # HOJA DESTINO
        # ==========================================
        ws = ws_diff if formula["tiene_diferencias"] else ws_ok

        # Título de la fórmula
        ws.append([f"{formula.get('idart', '')} - {formula['descripcion']}"])
        ws.cell(row=ws.max_row, column=1).font = BOLD

        # Encabezado de columnas
        encabezado = ["Insumo", "Unidad base", "Almacén base"] + [
            str(t) for t in columnas_tiendas
        ]
        ws.append(encabezado)
        row_header = ws.max_row

        for col in range(1, len(encabezado) + 1):
            c = ws.cell(row=row_header, column=col)
            c.font = BOLD
            c.fill = HEADER_FILL
            c.alignment = CENTER

        # Marcar encabezados de tiendas base con color diferenciado
        for i, tda in enumerate(columnas_tiendas):
            if int(tda) in bases_globales:
                c = ws.cell(row=row_header, column=4 + i)
                c.fill = BASE_FILL

        # ==========================================
        # FILAS DE INSUMOS
        # ==========================================
        for ins in formula["insumos"]:

            fila = [
                f"{ins.get('idinsumo', '')} - {ins['nombre']}",
                ins["unidad"],
                ins.get("almacen_base", "")
            ]

            for tda in columnas_tiendas:
                val = ins["tiendas"].get(tda)

                if val and val["cantidad"] is not None:
                    if val["diff"]:
                        # Ajuste 1: mostrar cantidad | unidad | almacen en celdas con diff
                        txt = (
                            f"{val['cantidad']:.2f}"
                            f" | {val.get('unidad', '')}"
                            f" | {val.get('almacen', '')}"
                        )
                    else:
                        txt = f"{val['cantidad']:.2f}"
                else:
                    txt = "-"

                fila.append(txt)

            ws.append(fila)
            row = ws.max_row

            # Resaltar celdas con diff en rojo
            for i, tda in enumerate(columnas_tiendas):
                val = ins["tiendas"].get(tda)
                if val and val["diff"]:
                    cell = ws.cell(row=row, column=4 + i)
                    cell.fill = DIFF_FILL
                    cell.font = WHITE_FONT

        ws.append([])

    for ws in wb.worksheets:
        ws.column_dimensions["A"].width = 37
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 5

        # # calcular máximo número de columnas usadas
        # max_cols = ws.max_column

        # for col in range(5, max_cols + 1):
        #     col_letter = get_column_letter(col)
        #     ws.column_dimensions[col_letter].width = 14
    # ==========================================
    # GUARDAR
    # ==========================================
    if not ruta_salida:
        ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        ruta_salida = f"SincroSIG/logs/formulaciones/reporte_formulaciones_{ts}.xlsx"

    os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
    try:
        wb.save(ruta_salida)
        print("EXCEL GUARDADO EN:", ruta_salida)
    except Exception as e:
        raise Exception(f"Error guardando Excel: {e}")

    return ruta_salida